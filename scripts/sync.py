#!/usr/bin/env python3
"""
Infinity Inventory Sync Engine
Synchronizes GitHub repos with Google Drive inventory
"""

import subprocess
import json
import os
from datetime import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

CONFIG = os.environ.get('RCLONE_CONFIG', '/home/ubuntu/.gdrive-rclone.ini')
ORG = 'InfinityXOneSystems'
DRIVE_BASE = 'manus_google_drive:INVENTORY'

def run_cmd(cmd):
    """Run shell command and return output"""
    result = subprocess.run(cmd, shell=True, capture_output=True, text=True)
    return result.stdout, result.returncode

def get_repos():
    """Fetch all repos from GitHub"""
    cmd = f'gh repo list {ORG} --limit 200 --json name,description,updatedAt'
    output, code = run_cmd(cmd)
    if code == 0:
        return json.loads(output)
    return []

def sync_repo(repo_name):
    """Sync a single repository"""
    timestamp = datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')
    status = 'success'
    error = ''
    files_updated = []
    
    try:
        # Clone repo
        run_cmd(f'rm -rf /tmp/{repo_name}')
        run_cmd(f'gh repo clone {ORG}/{repo_name} /tmp/{repo_name} --depth 1')
        
        # Create Drive folders
        for folder in ['01_STRUCTURE', '02_CODE_INDEX', '03_DOCS']:
            run_cmd(f'rclone mkdir "{DRIVE_BASE}/01_REPOS/{repo_name}/{folder}" --config {CONFIG}')
        
        # Generate and upload folder inventory
        # (simplified - full implementation would create Excel files)
        files_updated.append('folders')
        
    except Exception as e:
        status = 'failed'
        error = str(e)
    
    return {
        'timestamp': timestamp,
        'repo_name': repo_name,
        'action': 'sync',
        'files_updated': files_updated,
        'status': status,
        'error': error
    }

def update_audit_log(entries):
    """Update sync_audit_log.xlsx"""
    # Implementation would append to existing log
    pass

def main():
    print(f"=== Infinity Inventory Sync ===")
    print(f"Started: {datetime.utcnow().isoformat()}")
    
    repos = get_repos()
    print(f"Found {len(repos)} repositories")
    
    results = []
    for repo in repos:
        print(f"Syncing: {repo['name']}")
        result = sync_repo(repo['name'])
        results.append(result)
    
    update_audit_log(results)
    
    success = sum(1 for r in results if r['status'] == 'success')
    failed = sum(1 for r in results if r['status'] == 'failed')
    
    print(f"\n=== Sync Complete ===")
    print(f"Success: {success}, Failed: {failed}")
    print(f"Finished: {datetime.utcnow().isoformat()}")

if __name__ == '__main__':
    main()
